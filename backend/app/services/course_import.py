from __future__ import annotations

from datetime import date
from io import BytesIO

from openpyxl import load_workbook
from pydantic import ValidationError

from app.models.schedule import CourseImportResponse, CourseInput, DateRange, ScheduleProject, ValidationIssue
from app.services.validation import validate_project

EXPECTED_HEADERS = [
    "course id",
    "course name",
    "semester",
    "is high failure",
    "prerequisites",
]

TRUTHY_VALUES = {"1", "true", "yes", "y"}
FALSY_VALUES = {"0", "false", "no", "n", ""}


class CourseImportError(Exception):
    def __init__(self, issues: list[ValidationIssue]):
        super().__init__(issues[0].message if issues else "Course import failed.")
        self.issues = issues


def _normalize_header(value: object) -> str:
    return " ".join(str(value or "").strip().lower().split())


def _cell_text(value: object) -> str:
    return str(value or "").strip()


def _build_issue(message: str, row_number: int | None = None, course_code: str | None = None) -> ValidationIssue:
    row_prefix = f"Row {row_number}: " if row_number is not None else ""
    return ValidationIssue(
        code="invalid_import_file",
        severity="error",
        message=f"{row_prefix}{message}",
        related_course_code=course_code or None,
    )


def _parse_semester(value: object) -> int:
    if isinstance(value, bool):
        raise ValueError("Semester must be a number between 1 and 12.")
    if isinstance(value, int):
        return value
    if isinstance(value, float) and value.is_integer():
        return int(value)

    text = _cell_text(value)
    if not text:
        raise ValueError("Semester is required.")

    try:
        return int(text)
    except ValueError as error:
        raise ValueError("Semester must be a number between 1 and 12.") from error


def _parse_high_failure(value: object) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, int) and value in {0, 1}:
        return bool(value)
    if isinstance(value, float) and value in {0.0, 1.0}:
        return bool(int(value))

    normalized = _cell_text(value).lower()
    if normalized in TRUTHY_VALUES:
        return True
    if normalized in FALSY_VALUES:
        return False

    raise ValueError("Is High Failure must be one of yes/no, true/false, or 1/0.")


def import_courses_from_excel(content: bytes) -> CourseImportResponse:
    try:
        workbook = load_workbook(filename=BytesIO(content), data_only=True)
    except Exception as error:  # pragma: no cover - exact openpyxl error type is not stable
        raise CourseImportError([_build_issue("Uploaded file is not a valid .xlsx workbook.")]) from error

    worksheet = workbook.worksheets[0]
    actual_headers = [_normalize_header(worksheet.cell(row=1, column=column).value) for column in range(1, 6)]

    if actual_headers != EXPECTED_HEADERS:
        raise CourseImportError(
            [
                _build_issue(
                    "Template headers must be A: Course ID, B: Course Name, C: Semester, D: Is High Failure, E: Prerequisites.",
                )
            ]
        )

    issues: list[ValidationIssue] = []
    courses: list[CourseInput] = []

    for row_number in range(2, worksheet.max_row + 1):
        raw_values = [worksheet.cell(row=row_number, column=column).value for column in range(1, 6)]
        if all(_cell_text(value) == "" for value in raw_values):
            continue

        course_code = _cell_text(raw_values[0])
        course_name = _cell_text(raw_values[1])
        prerequisite_code = _cell_text(raw_values[4]) or None

        try:
            semester_number = _parse_semester(raw_values[2])
            high_failure_rate = _parse_high_failure(raw_values[3])
            courses.append(
                CourseInput(
                    course_code=course_code,
                    course_name=course_name,
                    semester_number=semester_number,
                    high_failure_rate=high_failure_rate,
                    prerequisite_course_code=prerequisite_code,
                )
            )
        except ValueError as error:
            issues.append(_build_issue(str(error), row_number=row_number, course_code=course_code or None))
        except ValidationError as error:
            for validation_error in error.errors():
                issues.append(_build_issue(validation_error["msg"], row_number=row_number, course_code=course_code or None))

    if not courses and not issues:
        issues.append(_build_issue("The workbook does not contain any course rows."))

    if issues:
        raise CourseImportError(issues)

    validation_project = ScheduleProject(
        project_name="Imported Courses",
        moed_a_window=DateRange(start_date=date(2026, 1, 1), end_date=date(2026, 1, 1)),
        courses=courses,
    )
    validation_issues = validate_project(validation_project)

    if validation_issues:
        raise CourseImportError(validation_issues)

    return CourseImportResponse(imported_count=len(courses), courses=courses)